from openpyxl import load_workbook

# Load workbook
loc = r"C:\Sachin\Programs\Python\Practice\sample1.xlsx"
wb = load_workbook(loc)

# Select the first sheet
sheet = wb.active  # or: wb.worksheets[0]

# # Accessing a specific cell
print(sheet.cell(row=3, column=2).value)   # Equivalent to cell_value(0, 0) in xlrd

# # Row values (row 2 and row 1)
print([cell.value for cell in sheet[2]])   # row index 1 in xlrd
print([cell.value for cell in sheet[1]])   # row index 0 in xlrd

# # Column values (column 3 and column 1)
print([sheet.cell(row=i, column=3).value for i in range(1, sheet.max_row + 1)])  # col index 2
print([sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)])  # col index 0
